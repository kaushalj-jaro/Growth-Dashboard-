"""
process_reports.py
-------------------
This is your original Growth Funnel Report notebook logic, unchanged,
wrapped in a single function so an upload portal (app.py) can call it
directly whenever new files are uploaded. Nothing about the calculations,
buckets, or dashboard construction has been touched -- only the file
paths at the very top now come in as arguments instead of being
hardcoded, and the very last step writes to Neon Postgres instead of
(in addition to) an .xlsx file.
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from db_writer import save_report_to_postgres

try:
    # Jupyter provides this automatically; outside a notebook it doesn't exist.
    display  # noqa: F821
except NameError:
    def display(x):
        print(x)


def run_pipeline(raw_path: str, enrolled_path: str, targets_path: str):
    """
    Runs the full Growth Funnel Report pipeline end to end:
      1. Loads the raw + enrolled CSVs and the targets workbook
      2. Runs every processing/aggregation step exactly as the notebook did
      3. Writes every output dashboard to Neon Postgres
      4. Returns the upload_id logged in the `upload_history` table

    raw_path, enrolled_path : paths to the two CSVs for this cycle
    targets_path            : path to the targets .xlsx for this cycle
    """

    # ============================================================
    # STEP 1 — CONFIG + GLOBAL SETUP (FORCED CYCLICAL OVERRIDE)
    # ============================================================

    import pandas as pd
    import numpy as np
    import re
    import os
    import calendar
    from openpyxl import load_workbook
    from datetime import datetime

    # ============================================================
    # DISPLAY SETTINGS
    # ============================================================

    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 5000)
    pd.set_option('display.max_colwidth', None)

    # ============================================================
    # FILE PATHS — MASTER FILES (PV Raw & PV Enrolled & TARGETS)
    # ============================================================
    # These now come in as function arguments from the upload portal
    # (whatever file the user just uploaded), instead of being hardcoded.

    MASTER_ENROLLED_PATH = enrolled_path
    MASTER_RAW_PATH = raw_path
    june_file = targets_path

    # ============================================================
    # AUTOMATIC DATE CONFIG FROM FILE PATHS
    # ============================================================

    enrolled_filename = os.path.basename(MASTER_ENROLLED_PATH)
    date_match = re.search(r'(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})', enrolled_filename)

    if date_match:
        file_day = int(date_match.group(1))
        file_month = date_match.group(2).capitalize()
        file_year = int(date_match.group(3))

        # Standard dynamic calendar assignment for the active month
        CYCLE_MONTH = file_month
        CYCLE_YEAR = file_year
        YESTERDAY_DAY = file_day
    else:
        # UPDATED: Safe manual fallback anchor now rolls over to June 2026
        CYCLE_MONTH = 'July'
        CYCLE_YEAR = 2026
        YESTERDAY_DAY = 2

    # Absolute Calendar Math Parameters Anchor
    month_idx = list(calendar.month_name).index(CYCLE_MONTH)
    DAYS_IN_MONTH = calendar.monthrange(CYCLE_YEAR, month_idx)[1]
    CURRENT_DAY = YESTERDAY_DAY + 1 if YESTERDAY_DAY < DAYS_IN_MONTH else YESTERDAY_DAY
    REMAINING_DAYS = max(DAYS_IN_MONTH - YESTERDAY_DAY, 0)

    CYCLE_START = pd.to_datetime(f'{CYCLE_YEAR}-{month_idx:02d}-01 00:00:00')

    # ============================================================
    # STAGE BUCKETS
    # ============================================================

    JUNK = ['Invalid', 'Disqualified', 'DND', 'Lead Not Enquired']
    WORKABLE = ['Callback', 'Eligible for DLP', 'Future Follow Up', 'Marketing Qualified', 'Not Contactable', 'Reborn',
                'Attempted']
    PROSPECT = ['Prospect', 'Meeting Done', 'Meeting Scheduled', 'Registered', 'Registered Lead', 'Opportunity']
    FRESH = ['Fresh']

    # ============================================================
    # SOURCE BUCKETS
    # ============================================================

    SOURCES = ['PR', 'WhatsApp', 'AI Meeting', 'Website', 'Live Chat', 'RDMPL']

    # ============================================================
    # PV UNIVERSITY BUCKET
    # ============================================================

    PV_UNIVERSITIES = [
        'IIM Ahmedabad', 'IIM Bangalore', 'IIM Calcutta', 'IIM Indore', 'IIM Kozhikode',
        'IIM Lucknow', 'IIM Mumbai', 'IIM Nagpur', 'IIM Raipur', 'IIM Ranchi',
        'IIM Rohtak', 'IIM Sambalpur', 'IIM Shillong', 'IIM Sirmaur', 'IIM Tiruchirappalli',
        'IIM Trichy', 'IIM Udaipur', 'IIM Visakhapatnam', 'IIT Bombay', 'IIT Delhi',
        'IIT Guwahati', 'IIT Kanpur', 'IIT Kharagpur', 'IIT Madras', 'IIT Roorkee', 'XLRI, Jamshedpur'
    ]
    PV_UNIVERSITIES = [str(x).strip() for x in PV_UNIVERSITIES]

    # ============================================================
    # FREE COURSE UNIVERSITY
    # ============================================================

    FREE_COURSE_UNIVERSITY = ['Jaro Education']

    # ============================================================
    # PRINT VALIDATION CONSOLE REPORTS
    # ============================================================

    print("\n===================================================")
    print("DATE CONFIGURATION RUNTIME VALIDATION")
    print("===================================================\n")
    print('Cycle Month    :', CYCLE_MONTH)
    print('Cycle Year     :', CYCLE_YEAR)
    print('Cycle Start    :', CYCLE_START)
    print('Current Day    :', CURRENT_DAY)
    print('Yesterday Day  :', YESTERDAY_DAY)
    print('Days In Month  :', DAYS_IN_MONTH)
    print('Remaining Days :', REMAINING_DAYS)

    print("\n===================================================")
    print("FILE PATHS MATRIX VALIDATION")
    print("===================================================\n")
    print("MASTER ENROLLED FILE :", MASTER_ENROLLED_PATH)
    print("MASTER RAW FILE      :", MASTER_RAW_PATH)
    print("TARGET FILE          :", june_file)

    print("\n✅ STEP 1 READY FOR RESILIENT RUNTIME EXECUTION")

    # ============================================================
    # STEP 2 — LOAD ALL DATASETS
    # Both DLP and PV are pulled from the master PV files.
    # ============================================================

    import pandas as pd

    # ============================================================
    # COMMON TEXT COLUMNS
    # ============================================================

    COMMON_TEXT_COLS = [
        'Branch',
        'University',
        'OFS 1 - Primary University',
        'Opportunity Source',
        'Stage',
        'Product',
        'Admission Branch',
        'Opportunity Id'
    ]

    # ============================================================
    # HELPER FUNCTION — CLEAN DATAFRAME
    # ============================================================

    def clean_dataframe(df):

        # --------------------------------------------------------
        # CLEAN COLUMN NAMES
        # --------------------------------------------------------

        df.columns = (
            df.columns
            .str.strip()
        )

        # --------------------------------------------------------
        # CLEAN TEXT COLUMNS
        # --------------------------------------------------------

        for col in COMMON_TEXT_COLS:

            if col in df.columns:
                df[col] = (
                    df[col]
                    .fillna('')
                    .astype(str)
                    .str.strip()
                    .str.replace(r'\s+', ' ', regex=True)
                )

        # --------------------------------------------------------
        # CLEAN OPPORTUNITY ID
        # --------------------------------------------------------

        if 'Opportunity Id' in df.columns:
            df['Opportunity Id'] = (
                df['Opportunity Id']
                .astype(str)
                .str.strip()
            )

        # --------------------------------------------------------
        # PARSE DATE
        # --------------------------------------------------------

        if 'Created On' in df.columns:
            df['Created On'] = pd.to_datetime(
                df['Created On'],
                errors='coerce'
            )

        return df

    # ============================================================
    # LOAD MASTER RAW
    # (Contains both DLP + PV + Free Courses)
    # ============================================================

    master_raw = pd.read_csv(
        MASTER_RAW_PATH,
        low_memory=False
    )

    master_raw = clean_dataframe(master_raw)

    print("✅ MASTER RAW LOADED")
    print("Rows :", len(master_raw))
    print()

    # ============================================================
    # LOAD MASTER ENROLLED
    # (Contains both DLP + PV + Free Courses)
    # ============================================================

    master_enrolled = pd.read_csv(
        MASTER_ENROLLED_PATH,
        low_memory=False
    )

    master_enrolled = clean_dataframe(master_enrolled)

    print("✅ MASTER ENROLLED LOADED")
    print("Rows :", len(master_enrolled))
    print()

    # ============================================================
    # ASSIGN SEGMENT — RAW
    # ============================================================

    master_raw['Segment'] = 'DLP'

    pv_mask = (

            master_raw['University'].isin(PV_UNIVERSITIES)

            |

            master_raw['OFS 1 - Primary University'].isin(PV_UNIVERSITIES)

    )

    master_raw.loc[
        pv_mask,
        'Segment'
    ] = 'PV'

    master_raw.loc[
        master_raw['University'] == 'Jaro Education',
        'Segment'
    ] = 'Free Courses'

    master_raw['Data Source'] = 'MASTER'

    # ============================================================
    # FREE COURSE IDS — MAP TO ENROLLED
    # ============================================================

    free_course_ids = set(
        master_raw[
            master_raw['Segment'] == 'Free Courses'
            ]['Opportunity Id']
    )

    # ============================================================
    # ASSIGN SEGMENT — ENROLLED
    # ============================================================

    master_enrolled['Segment'] = 'DLP'

    pv_mask_enrolled = (

            master_enrolled['University'].isin(PV_UNIVERSITIES)

            |

            master_enrolled['OFS 1 - Primary University'].isin(PV_UNIVERSITIES)

    )

    master_enrolled.loc[
        pv_mask_enrolled,
        'Segment'
    ] = 'PV'

    master_enrolled.loc[
        master_enrolled['Opportunity Id'].isin(free_course_ids),
        'Segment'
    ] = 'Free Courses'

    master_enrolled['Data Source'] = 'MASTER'

    # ============================================================
    # CREATE SEGMENT VIEWS FOR COMPATIBILITY
    # (Replaces the old dlp_raw / pv_raw / dlp_enrolled / pv_enrolled)
    # ============================================================

    dlp_raw = master_raw[master_raw['Segment'] == 'DLP'].copy()
    pv_raw = master_raw[master_raw['Segment'].isin(['PV', 'Free Courses'])].copy()

    dlp_enrolled = master_enrolled[master_enrolled['Segment'] == 'DLP'].copy()
    pv_enrolled = master_enrolled[master_enrolled['Segment'].isin(['PV', 'Free Courses'])].copy()

    # ============================================================
    # SEGMENT VALIDATION — RAW
    # ============================================================

    print("\n===================================================")
    print("MASTER RAW VALIDATION")
    print("===================================================\n")

    print("Rows :", len(master_raw))

    print("\nColumns:\n")
    print(master_raw.columns.tolist())

    # ============================================================
    # SEGMENT VALIDATION — ENROLLED
    # ============================================================

    print("\n===================================================")
    print("MASTER ENROLLED VALIDATION")
    print("===================================================\n")

    print("Rows :", len(master_enrolled))

    print("\nColumns:\n")
    print(master_enrolled.columns.tolist())

    # ============================================================
    # SEGMENT SPLIT VALIDATION
    # ============================================================

    print("\n===================================================")
    print("SEGMENT SPLIT — RAW")
    print("===================================================\n")

    print(
        master_raw['Segment']
        .value_counts(dropna=False)
    )

    print("\n===================================================")
    print("SEGMENT SPLIT — ENROLLED")
    print("===================================================\n")

    print(
        master_enrolled['Segment']
        .value_counts(dropna=False)
    )

    # ============================================================
    # FREE COURSE VALIDATION
    # ============================================================

    print("\n===================================================")
    print("FREE COURSE VALIDATION")
    print("===================================================\n")

    print(
        "FREE COURSE RAW ROWS :",
        len(master_raw[master_raw['Segment'] == 'Free Courses'])
    )

    print()

    print(
        "FREE COURSE ENROLLED ROWS :",
        len(master_enrolled[master_enrolled['Segment'] == 'Free Courses'])
    )

    # ============================================================
    # TOTAL VALIDATION
    # ============================================================

    print("\n===================================================")
    print("TOTAL DATA VALIDATION")
    print("===================================================\n")

    print("TOTAL RAW ROWS     :", len(master_raw))
    print("TOTAL ENROLLED ROWS :", len(master_enrolled))

    # ============================================================
    # FINAL READY
    # ============================================================

    print("\n✅ STEP 2 READY")

    # ============================================================
    # STEP 3 — MASTER CLEANING & STANDARDIZATION
    # ============================================================

    import pandas as pd
    import numpy as np

    # ============================================================
    # COMMON UNIVERSITY STANDARDIZATION
    # ============================================================

    UNIVERSITY_STANDARDIZATION = {

        # --------------------------------------------------------
        # SYMBIOSIS
        # --------------------------------------------------------

        'Symbiosis School for Online and Digital Learning':
            'Symbiosis School For Online And Digital Learning',

        # --------------------------------------------------------
        # MUJ
        # --------------------------------------------------------

        'Manipal Unniversity, Jaipur':
            'Manipal University, Jaipur',

        'MUJ':
            'Manipal University, Jaipur',

        # --------------------------------------------------------
        # MGR
        # --------------------------------------------------------

        'MGR':
            'Dr. M.G.R. University',

        # --------------------------------------------------------
        # AMRITA
        # --------------------------------------------------------

        'Amrita':
            'Amrita Vishwa Vidyapeetham',

        # --------------------------------------------------------
        # IIM
        # --------------------------------------------------------

        'IIM Online':
            'IIM',

    }

    # ============================================================
    # COMMON BRANCH STANDARDIZATION
    # ============================================================

    BRANCH_STANDARDIZATION = {

        'Pune1': 'Pune 1',

        'Thane 1': 'Thane',
        'Noida': 'Noida 1',
        'Pune': 'Pune 1',
        'Thane 1': 'Thane',
        'Thane 2': 'Thane',
        'Goregoan': 'Goregaon',

    }

    # ============================================================
    # COMMON SOURCE STANDARDIZATION
    # ============================================================

    SOURCE_STANDARDIZATION = {

        # --------------------------------------------------------
        # WHATSAPP
        # --------------------------------------------------------

        'whatsapp': 'WhatsApp',
        'whatsapp reference': 'WhatsApp',
        'whatsapp ref': 'WhatsApp',
        'whatsapp lead': 'WhatsApp',
        'whatsapp leads': 'WhatsApp',
        'whatsappreference': 'WhatsApp',
        'whats app': 'WhatsApp',
        'wa': 'WhatsApp',
        'wa ref': 'WhatsApp',
        'wa reference': 'WhatsApp',

        # --------------------------------------------------------
        # PR
        # --------------------------------------------------------

        'pr': 'PR',
        'prchat': 'PR',
        'pr story': 'PR',
        'prstory': 'PR',
        'prstory campaign': 'PR',

        # --------------------------------------------------------
        # AI MEETING
        # --------------------------------------------------------

        'ai meeting': 'AI Meeting',
        'career counselling': 'AI Meeting',
        'career counseling': 'AI Meeting',
        'superbot': 'AI Meeting',

        # --------------------------------------------------------
        # WEBSITE
        # --------------------------------------------------------

        'website': 'Website',
        'inbound call': 'Website',
        'inbound calls': 'Website',
        'inbound phone call': 'Website',
        'inbound phone calls': 'Website',

        # --------------------------------------------------------
        # LIVE CHAT
        # --------------------------------------------------------

        'live chat': 'Live Chat',
        'livechat': 'Live Chat',
        'live_chat': 'Live Chat',
        'chat': 'Live Chat',
        'live-chat': 'Live Chat',
        'lc': 'Live Chat',

        'rdmpl': 'RDMPL',
        'rdmpl lead': 'RDMPL',
        'rdmpl campaign': 'RDMPL',

    }

    # ============================================================
    # MASTER CLEAN FUNCTION
    # ============================================================

    def clean_dataset(
            df,
            dataset_name='DATASET',
            is_enrolled=False
    ):

        print("\n===================================================")
        print(f"CLEANING : {dataset_name}")
        print("===================================================\n")

        # 1. Clean Column Headers
        df.columns = df.columns.str.strip()

        # 2. Standardize Ingress Timestamps
        if 'Created On' in df.columns:
            df['Created On'] = pd.to_datetime(
                df['Created On'],
                dayfirst=True,
                errors='coerce'
            )

        text_cols = [
            'Branch',
            'Admission Branch',
            'University',
            'OFS 1 - Primary University',
            'Opportunity Source',
            'Stage',
            'DER Month',
            'Product',
            'Opportunity Id'
        ]

        # 3. UPGRADED: Clean and Strip Whitespace / Escape Tabs and Newlines
        for col in text_cols:
            if col in df.columns:
                df[col] = (
                    df[col]
                    .fillna('')
                    .astype(str)
                    .str.strip()
                    # Safely escape carriage/newline layout elements first
                    .str.replace(r'[\n\r\t]', ' ', regex=True)
                    # Compress any residual duplicate spaces down to single spaces
                    .str.replace(r'\s+', ' ', regex=True)
                )

        # 4. Apply Geographic Branch Normalization Maps
        if 'Branch' in df.columns:
            df['Branch'] = df['Branch'].replace(BRANCH_STANDARDIZATION)

        if 'Admission Branch' in df.columns:
            df['Admission Branch'] = df['Admission Branch'].replace(BRANCH_STANDARDIZATION)

        # 5. Apply Academic Partner Normalization Maps
        if 'University' in df.columns:
            df['University'] = df['University'].replace(UNIVERSITY_STANDARDIZATION)

        if 'OFS 1 - Primary University' in df.columns:
            df['OFS 1 - Primary University'] = df['OFS 1 - Primary University'].replace(UNIVERSITY_STANDARDIZATION)

        # 6. Apply Standard Source Groupings
        if 'Opportunity Source' in df.columns:
            df['Opportunity Source'] = (
                df['Opportunity Source']
                .fillna('')
                .astype(str)
                .str.strip()
                .str.lower()
                .replace(SOURCE_STANDARDIZATION)
            )

        # 7. Enforce Global Branch Cleaning Rule
        if 'Branch' in df.columns:
            df = df[
                (df['Branch'] != '') &
                (df['Branch'].str.lower() != 'nan') &
                (df['Branch'].isna() == False)
                ].copy()

        # Additional protection for Enrolled sheets
        if is_enrolled and 'Admission Branch' in df.columns:
            df = df[
                (df['Admission Branch'] != '') &
                (df['Admission Branch'].str.lower() != 'nan') &
                (df['Admission Branch'].isna() == False)
                ].copy()

        # 8. Run Master Deduplication Key Checks
        if 'Opportunity Id' in df.columns:
            df = df.drop_duplicates(subset=['Opportunity Id'])

        print("Rows :", len(df))

        print("\nBranch Count :")
        if 'Branch' in df.columns:
            print(df['Branch'].nunique())

        print("\nUniversity Count :")
        if 'University' in df.columns:
            print(df['University'].nunique())

        print("\nSource Distribution :")
        if 'Opportunity Source' in df.columns:
            print(
                df['Opportunity Source']
                .value_counts(dropna=False)
                .to_string()
            )

        return df

    # ============================================================
    # OPERATIONAL INGESTION RUNS
    # ============================================================

    # Clean Master Raw Funnel Events
    master_raw = clean_dataset(
        master_raw,
        dataset_name='MASTER RAW',
        is_enrolled=False
    )

    # Clean Master Enrolled Cohort Events
    master_enrolled = clean_dataset(
        master_enrolled,
        dataset_name='MASTER ENROLLED',
        is_enrolled=True
    )

    # ============================================================
    # RE-APPLY SEGMENT AFTER STANDARDIZATION
    # ============================================================

    master_raw['Segment'] = 'DLP'

    pv_mask = (
            master_raw['University'].isin(PV_UNIVERSITIES)
            |
            master_raw['OFS 1 - Primary University'].isin(PV_UNIVERSITIES)
    )

    master_raw.loc[pv_mask, 'Segment'] = 'PV'

    master_raw.loc[
        master_raw['University'] == 'Jaro Education',
        'Segment'
    ] = 'Free Courses'

    # ============================================================
    # FREE COURSE IDS — REFRESH AFTER STANDARDIZATION
    # ============================================================

    free_course_ids = set(
        master_raw[
            master_raw['Segment'] == 'Free Courses'
            ]['Opportunity Id']
    )

    master_enrolled['Segment'] = 'DLP'

    master_enrolled.loc[
        master_enrolled['University'].isin(PV_UNIVERSITIES),
        'Segment'
    ] = 'PV'

    master_enrolled.loc[
        master_enrolled['Opportunity Id'].isin(free_course_ids),
        'Segment'
    ] = 'Free Courses'

    # ============================================================
    # REFRESH SEGMENT VIEWS
    # ============================================================

    dlp_raw = master_raw[master_raw['Segment'] == 'DLP'].copy()
    pv_raw = master_raw[master_raw['Segment'].isin(['PV', 'Free Courses'])].copy()

    dlp_enrolled = master_enrolled[master_enrolled['Segment'] == 'DLP'].copy()
    pv_enrolled = master_enrolled[master_enrolled['Segment'].isin(['PV', 'Free Courses'])].copy()

    # ============================================================
    # ADD XLRI TO PV IF MISSING
    # ============================================================

    if 'XLRI, Jamshedpur' not in PV_UNIVERSITIES:
        PV_UNIVERSITIES.append('XLRI, Jamshedpur')

    # ============================================================
    # FINAL RECONCILIATION VALIDATION PRINT-BLOCK
    # ============================================================

    print("\n===================================================")
    print("FINAL SEGMENT VALIDATION — RAW")
    print("===================================================\n")

    print(master_raw['Segment'].value_counts(dropna=False))

    print("\n===================================================")
    print("FINAL SEGMENT VALIDATION — ENROLLED")
    print("===================================================\n")

    print(master_enrolled['Segment'].value_counts(dropna=False))

    print("\n===================================================")
    print("FINAL DATASET COUNTS")
    print("===================================================\n")

    print("MASTER RAW Rows       :", len(master_raw))
    print("MASTER ENROLLED Rows  :", len(master_enrolled))
    print()
    print("DLP RAW Rows          :", len(dlp_raw))
    print("DLP ENROLLED Rows     :", len(dlp_enrolled))
    print("PV RAW Rows           :", len(pv_raw))
    print("PV ENROLLED Rows      :", len(pv_enrolled))

    print("\n===================================================")
    print("PV UNIVERSITY VALIDATION")
    print("===================================================\n")

    print("PV RAW Universities:\n")
    print(sorted(pv_raw['University'].dropna().unique()))

    print("\nPV ENROLLED Universities:\n")
    print(sorted(pv_enrolled['University'].dropna().unique()))

    print("\n===================================================")
    print("FREE COURSE VALIDATION")
    print("===================================================\n")

    print(
        "FREE COURSE RAW ROWS :",
        len(master_raw[master_raw['Segment'] == 'Free Courses'])
    )

    print()

    print(
        "FREE COURSE ENROLLED ROWS :",
        len(master_enrolled[master_enrolled['Segment'] == 'Free Courses'])
    )

    print("\n✅ STEP 3 READY")

    # ============================================================
    # STEP 4 — TARGET EXTRACTION (DLP + PV + FREE COURSES)
    # ============================================================

    from openpyxl import load_workbook
    import pandas as pd

    # ============================================================
    # LOAD WORKBOOK
    # ============================================================

    wb = load_workbook(
        june_file,
        read_only=True
    )

    # ============================================================
    # HELPER FUNCTION
    # ============================================================

    def parse_cell(val):

        if val is None:
            return 0

        cleaned = (
            str(val)
            .strip()
            .replace(',', '')
            .replace('=', '')
        )

        try:
            return int(float(cleaned))
        except:
            return 0

    # ============================================================
    # BRANCH FIXES
    # ============================================================

    BRANCH_FIX = {

        'Pune1': 'Pune 1',

        'Thane 1': 'Thane',
        'Noida': 'Noida 1',
        'Pune': 'Pune 1',
        'Thane 1': 'Thane',
        'Thane 2': 'Thane',
        'Goregoan': 'Goregaon',

    }

    # ============================================================
    # BRANCH -> VP MAPPING (fixed constant, doesn't change month to
    # month -- edit this dict directly if a branch changes ownership)
    # ============================================================

    BRANCH_VP_MAP = {
        'Ahmedabad': 'Viral Sir',
        'Bangalore': 'Anand Sir',
        'Chandigarh': 'Anand Sir',
        'Chembur': 'Unknown VP',  # TODO: confirm correct VP
        'Chembur HO': 'Anand Sir',
        'Chennai': 'Viral Sir',
        'Dadar': 'Anand Sir',
        'Goregaon': 'Viral Sir',
        'Gurugram': 'Anand Sir',
        'Hyderabad': 'Nihal Sir',
        'Indore': 'Viral Sir',
        'Jaipur': 'Anand Sir',
        'Kochi': 'Nihal Sir',
        'Kolkata': 'Viral Sir',
        'Lucknow': 'Unknown VP',  # TODO: confirm correct VP
        'Noida 1': 'Nihal Sir',
        'Noida 2': 'Nihal Sir',
        'Pune 1': 'Nihal Sir',
        'Pune 2': 'Unknown VP',  # TODO: confirm correct VP
        'Sakinaka': 'Anand Sir',
        'Thane': 'Unknown VP',  # TODO: confirm correct VP
    }

    # ============================================================
    # UNIVERSITY FIXES
    # ============================================================

    SYMB_FIX = 'Symbiosis School For Online And Digital Learning'

    UNIVERSITY_FIX = {

        'SIU': SYMB_FIX,

        'BVDU': 'Bharati Vidyapeeth',

        'DPU': 'D Y Patil Vidyapeeth',

        'CU': 'Chandigarh University',

        'Amity': 'Amity University',

        'Amrita': 'Amrita Vishwa Vidyapeetham',

        'MGR': 'Dr. M.G.R. University',

        'MUJ': 'Manipal University, Jaipur',

        'Parul': 'Parul University',

        'Others': 'Any',

        'Manipal Unniversity, Jaipur':
            'Manipal University, Jaipur'

    }

    # ============================================================
    # BRANCH TARGET EXTRACTION
    # ============================================================

    branch_targets = {}
    branch_channel_targets = {}  # {branch: {'PR': x, 'AI': y, 'WA': z}}

    current_branch = None

    for row in wb['Branch-Product'].iter_rows(values_only=True):

        branch_cell = row[0]

        if branch_cell == 'Branch':
            continue

        if branch_cell is not None:

            current_branch = str(branch_cell).strip()

            current_branch = BRANCH_FIX.get(
                current_branch,
                current_branch
            )

            if current_branch not in branch_targets:
                branch_targets[current_branch] = 0

            if current_branch not in branch_channel_targets:
                branch_channel_targets[current_branch] = {'PR': 0, 'AI': 0, 'WA': 0}

        if current_branch:
            pr_target = parse_cell(row[2])
            ai_target = parse_cell(row[3])
            wa_target = parse_cell(row[4])

            branch_targets[current_branch] += (
                    pr_target + ai_target + wa_target
            )

            branch_channel_targets[current_branch]['PR'] += pr_target
            branch_channel_targets[current_branch]['AI'] += ai_target
            branch_channel_targets[current_branch]['WA'] += wa_target

    # ============================================================
    # PER-BRANCH PR/AI/WA TARGET DF (feeds Enhanced Branch Report)
    # Web has no source target anywhere in the targets workbook, so
    # Web Target / Web Till Date Target / Web Deficit stay as "-"
    # on every branch -- this is intentional, not a bug.
    # ============================================================

    df_branch_channel_targets = pd.DataFrame([
        {
            'Branch': b,
            'PR Target': vals['PR'],
            'AI Target': vals['AI'],
            'WA Target': vals['WA'],
        }
        for b, vals in branch_channel_targets.items()
    ])

    for ch in ['PR', 'AI', 'WA']:
        df_branch_channel_targets[f'{ch} Till Date Target'] = round(
            df_branch_channel_targets[f'{ch} Target']
            / DAYS_IN_MONTH
            * YESTERDAY_DAY
        ).astype(int)

    df_branch_channel_targets['Web Target'] = '-'
    df_branch_channel_targets['Web Till Date Target'] = '-'
    df_branch_channel_targets['Web Deficit'] = '-'

    # ============================================================
    # CREATE BRANCH TARGET DF
    # ============================================================

    df_branch_targets = pd.DataFrame(
        list(branch_targets.items()),
        columns=['Branch', 'Overall Target']
    )

    df_branch_targets = df_branch_targets[
        ~df_branch_targets['Branch'].isin([
            'Noida',
            'Noida-1&2',
            'Pune',
            'Pune-1 & 2'
        ])
    ]

    df_branch_targets['Till Date Target'] = round(
        df_branch_targets['Overall Target']
        / DAYS_IN_MONTH
        * YESTERDAY_DAY
    ).astype(int)

    df_branch_targets = (
        df_branch_targets
        .sort_values('Branch')
        .reset_index(drop=True)
    )

    branch_grand_total = pd.DataFrame([{
        'Branch': 'Grand Total',
        'Overall Target': df_branch_targets['Overall Target'].sum(),
        'Till Date Target': df_branch_targets['Till Date Target'].sum()
    }])

    df_branch_targets = pd.concat(
        [df_branch_targets, branch_grand_total],
        ignore_index=True
    )

    print("\n===================================================")
    print("BRANCH TARGETS")
    print("===================================================\n")

    display(df_branch_targets)

    # ============================================================
    # PR TARGET EXTRACTION
    # ============================================================

    pr_records = []

    for row in wb['Sheet7'].iter_rows(values_only=True):

        if row[0] in (None, 'Product', 'Total'):
            continue

        university = UNIVERSITY_FIX.get(
            str(row[0]).strip(),
            str(row[0]).strip()
        )

        pr_records.append({
            'University': university,
            'PR Leads Target': parse_cell(row[1])
        })

    df_pr_targets = pd.DataFrame(pr_records)

    # ============================================================
    # AI + WA TARGET EXTRACTION
    # ============================================================

    aiwa_records = []

    for row in wb['AI Meeting & Whatsapp Lead Targ'].iter_rows(values_only=True):

        if row[0] in (None, 'University', 'Total'):
            continue

        university = UNIVERSITY_FIX.get(
            str(row[0]).strip(),
            str(row[0]).strip()
        )

        aiwa_records.append({
            'University': university,
            'AI Leads Target': parse_cell(row[1]),
            'WhatsApp Leads Target': parse_cell(row[2])
        })

    df_aiwa_targets = pd.DataFrame(aiwa_records)

    # ============================================================
    # MERGE TARGETS
    # ============================================================

    uni_targets_df = df_pr_targets.merge(
        df_aiwa_targets,
        on='University',
        how='outer'
    ).fillna(0)

    uni_targets_df = uni_targets_df[
        uni_targets_df['University'] != 'Any'
        ]

    target_cols = [
        'PR Leads Target',
        'AI Leads Target',
        'WhatsApp Leads Target'
    ]

    for col in target_cols:
        uni_targets_df[col] = (
            pd.to_numeric(uni_targets_df[col], errors='coerce')
            .fillna(0)
            .astype(int)
        )

    uni_targets_df['Overall Target'] = (
            uni_targets_df['PR Leads Target']
            + uni_targets_df['AI Leads Target']
            + uni_targets_df['WhatsApp Leads Target']
    )

    uni_targets_df['Till Date Target'] = round(
        uni_targets_df['Overall Target']
        / DAYS_IN_MONTH
        * YESTERDAY_DAY
    ).astype(int)

    # ============================================================
    # DLP / PV / FREE COURSE TARGETS
    # ============================================================

    dlp_uni_targets = uni_targets_df[
        ~uni_targets_df['University']
        .isin(PV_UNIVERSITIES + FREE_COURSE_UNIVERSITY)
    ].copy()

    pv_uni_targets = uni_targets_df[
        uni_targets_df['University'].isin(PV_UNIVERSITIES)
    ].copy()

    free_course_targets = uni_targets_df[
        uni_targets_df['University'].isin(FREE_COURSE_UNIVERSITY)
    ].copy()

    dlp_uni_targets = dlp_uni_targets.sort_values('University').reset_index(drop=True)
    pv_uni_targets = pv_uni_targets.sort_values('University').reset_index(drop=True)
    free_course_targets = free_course_targets.sort_values('University').reset_index(drop=True)

    def add_grand_total(df):

        total_df = pd.DataFrame([{
            'University': 'Grand Total',
            'PR Leads Target': df['PR Leads Target'].sum(),
            'AI Leads Target': df['AI Leads Target'].sum(),
            'WhatsApp Leads Target': df['WhatsApp Leads Target'].sum(),
            'Overall Target': df['Overall Target'].sum(),
            'Till Date Target': df['Till Date Target'].sum()
        }])

        return pd.concat([df, total_df], ignore_index=True)

    dlp_uni_targets = add_grand_total(dlp_uni_targets)
    pv_uni_targets = add_grand_total(pv_uni_targets)
    free_course_targets = add_grand_total(free_course_targets)

    print("\n===================================================")
    print("DLP UNIVERSITY TARGETS")
    print("===================================================\n")
    display(dlp_uni_targets)

    print("\n===================================================")
    print("PV UNIVERSITY TARGETS")
    print("===================================================\n")
    display(pv_uni_targets)

    print("\n===================================================")
    print("FREE COURSE TARGETS")
    print("===================================================\n")
    display(free_course_targets)

    print("\n===================================================")
    print("TARGET VALIDATION")
    print("===================================================\n")

    print(
        "DLP Universities :",
        dlp_uni_targets[dlp_uni_targets['University'] != 'Grand Total']['University'].nunique()
    )

    print(
        "PV Universities :",
        pv_uni_targets[pv_uni_targets['University'] != 'Grand Total']['University'].nunique()
    )

    print(
        "Free Course Universities :",
        free_course_targets[free_course_targets['University'] != 'Grand Total']['University'].nunique()
    )

    print(
        "Branch Targets :",
        df_branch_targets[df_branch_targets['Branch'] != 'Grand Total']['Branch'].nunique()
    )

    print("\n✅ STEP 4 READY")

    # ============================================================
    # STEP 5 — FINAL PRODUCTION SOURCE STANDARDIZATION
    # ============================================================

    import pandas as pd

    # ============================================================
    # MASTER SOURCE STANDARDIZATION
    # ============================================================

    SOURCE_STANDARDIZATION = {

        # --------------------------------------------------------
        # WHATSAPP
        # --------------------------------------------------------

        'whatsapp': 'WhatsApp',
        'whatsapp reference': 'WhatsApp',
        'whatsapp ref': 'WhatsApp',
        'whatsapp lead': 'WhatsApp',
        'whatsapp leads': 'WhatsApp',
        'whatsappreference': 'WhatsApp',
        'whats app': 'WhatsApp',

        'wa': 'WhatsApp',
        'wa ref': 'WhatsApp',
        'wa reference': 'WhatsApp',

        # --------------------------------------------------------
        # PR
        # --------------------------------------------------------

        'pr': 'PR',
        'prchat': 'PR',
        'pr story': 'PR',
        'prstory': 'PR',
        'prstory campaign': 'PR',

        # --------------------------------------------------------
        # AI MEETING
        # --------------------------------------------------------

        'ai meeting': 'AI Meeting',
        'career counselling': 'AI Meeting',
        'career counseling': 'AI Meeting',
        'superbot': 'AI Meeting',

        # --------------------------------------------------------
        # WEBSITE
        # --------------------------------------------------------

        'website': 'Website',
        'inbound call': 'Website',
        'inbound calls': 'Website',
        'inbound phone call': 'Website',
        'inbound phone calls': 'Website',
        'inbound enquiry': 'Website',
        'inbound inquiry': 'Website',
        'web': 'Website',
        'organic website': 'Website',
        'website lead': 'Website',

        'rdmpl': 'RDMPL',
        'rdmpl lead': 'RDMPL',
        'rdmpl campaign': 'RDMPL',

    }

    # ============================================================
    # SOURCE CLEAN FUNCTION
    # ============================================================

    def standardize_sources(df, dataset_name='DATASET'):

        print("\n===================================================")
        print(f"SOURCE STANDARDIZATION : {dataset_name}")
        print("===================================================\n")

        df['Opportunity Source'] = (
            df['Opportunity Source']
            .fillna('')
            .astype(str)
            .str.strip()
            .str.lower()
        )

        df['Opportunity Source'] = (
            df['Opportunity Source']
            .replace(SOURCE_STANDARDIZATION)
        )

        df['Opportunity Source'] = (
            df['Opportunity Source']
            .replace('', 'Others')
        )

        # --------------------------------------------------------
        # FINAL NORMALISATION — fix any remaining lowercase leakage
        # e.g. if 'live chat' was not caught by replace above
        # --------------------------------------------------------

        FINAL_NORM = {
            'live chat': 'Live Chat',
            'livechat': 'Live Chat',
            'live-chat': 'Live Chat',
            'live_chat': 'Live Chat',
            'lc': 'Live Chat',
            'chat': 'Live Chat',
            'superbot': 'AI Meeting',
            'whatsapp': 'WhatsApp',
            'pr': 'PR',
            'prchat': 'PR',
            'website': 'Website',
            'ai meeting': 'AI Meeting',
            'rdmpl': 'RDMPL',

        }

        df['Opportunity Source'] = (
            df['Opportunity Source']
            .replace(FINAL_NORM)
        )

        validation = (
            df['Opportunity Source']
            .value_counts(dropna=False)
            .reset_index()
        )

        validation.columns = ['Source', 'Count']

        print(validation.to_string(index=False))

        return df

    # ============================================================
    # APPLY — MASTER RAW & ENROLLED
    # ============================================================

    master_raw = standardize_sources(master_raw, dataset_name='MASTER RAW')
    master_enrolled = standardize_sources(master_enrolled, dataset_name='MASTER ENROLLED')

    # ============================================================
    # REFRESH SEGMENT VIEWS
    # ============================================================

    dlp_raw = master_raw[master_raw['Segment'] == 'DLP'].copy()
    pv_raw = master_raw[master_raw['Segment'].isin(['PV', 'Free Courses'])].copy()

    dlp_enrolled = master_enrolled[master_enrolled['Segment'] == 'DLP'].copy()
    pv_enrolled = master_enrolled[master_enrolled['Segment'].isin(['PV', 'Free Courses'])].copy()

    # ============================================================
    # SOURCE MAP
    # ============================================================

    SOURCE_MAP = {
        'WhatsApp': ['WhatsApp'],
        'PR': ['PR'],
        'AI Meeting': ['AI Meeting'],
        'Website': ['Website'],
        'Live Chat': ['Live Chat'],
        'RDMPL': ['RDMPL'],
    }

    # ============================================================
    # VALIDATION FUNCTION
    # ============================================================

    def validate_sources(df, dataset_name='DATASET'):

        print("\n===================================================")
        print(f"FINAL SOURCE VALIDATION : {dataset_name}")
        print("===================================================\n")

        source_counts = df['Opportunity Source'].value_counts(dropna=False)
        print(source_counts)

        mapped_sources = ['PR', 'WhatsApp', 'AI Meeting', 'Website', 'Live Chat']

        mapped_total = len(df[df['Opportunity Source'].isin(mapped_sources)])
        total_rows = len(df)
        other_rows = total_rows - mapped_total

        print("\n---------------------------------------------------")
        print("TOTAL ROWS         :", total_rows)
        print("MAPPED MAIN SOURCE :", mapped_total)
        print("OTHER / UNMAPPED   :", other_rows)

        if other_rows > 0:
            print("\nOTHER SOURCE BREAKDOWN:\n")

            print(
                df[~df['Opportunity Source'].isin(mapped_sources)]
                ['Opportunity Source']
                .value_counts(dropna=False)
                .to_string()
            )

    validate_sources(master_raw, dataset_name='MASTER RAW')
    validate_sources(master_enrolled, dataset_name='MASTER ENROLLED')

    print("\n===================================================")
    print("SEGMENT + SOURCE VALIDATION")
    print("===================================================\n")

    print("MASTER RAW")
    print(pd.crosstab(master_raw['Segment'], master_raw['Opportunity Source']))

    print()

    print("MASTER ENROLLED")
    print(pd.crosstab(master_enrolled['Segment'], master_enrolled['Opportunity Source']))

    print("\n===================================================")
    print("MASTER DATA VALIDATION")
    print("===================================================\n")

    print("MASTER RAW Rows       :", len(master_raw))
    print("MASTER ENROLLED Rows  :", len(master_enrolled))
    print()
    print("DLP RAW Rows          :", len(dlp_raw))
    print("DLP ENROLLED Rows     :", len(dlp_enrolled))
    print("PV RAW Rows           :", len(pv_raw))
    print("PV ENROLLED Rows      :", len(pv_enrolled))
    print()
    print("TOTAL RAW ROWS        :", len(master_raw))
    print("TOTAL ENROLLED ROWS   :", len(master_enrolled))

    print("\n✅ STEP 5 READY")

    # ============================================================
    # STEP 6 — PRODUCTION MATRIX FILTER (ELIMINATES DUPLICATES)
    # ============================================================

    import pandas as pd
    import numpy as np

    # Create master datasets
    df_raw = pd.concat([dlp_raw, pv_raw], ignore_index=True).copy()
    df_enrolled = pd.concat([dlp_enrolled, pv_enrolled], ignore_index=True).copy()

    print("\n===================================================")
    print("MASTER DATA VALIDATION")
    print("===================================================\n")
    print("RAW ROWS (WITH LIVE CHAT)        :", len(df_raw))
    print("ENROLLED ROWS (WITH LIVE CHAT)   :", len(df_enrolled))

    # Required columns validation
    required_cols = ['Branch', 'University', 'Admission Branch', 'OFS 1 - Primary University', 'Segment',
                     'Opportunity Id', 'Opportunity Source']
    for col in required_cols:
        if col not in df_raw.columns: df_raw[col] = ''
        if col not in df_enrolled.columns: df_enrolled[col] = ''

    def clean_text(series):
        return series.fillna('').astype(str).str.strip().str.replace(r'[\n\r\t]', ' ', regex=True).str.replace(r'\s+',
                                                                                                               ' ',
                                                                                                               regex=True)

    for col in ['Branch', 'University', 'OFS 1 - Primary University', 'Segment', 'Opportunity Source']:
        df_raw[col] = clean_text(df_raw[col])

    for col in ['Branch', 'Admission Branch', 'University', 'Segment', 'Opportunity Source']:
        df_enrolled[col] = clean_text(df_enrolled[col])

    # University Standardization Cleanup
    UNIVERSITY_STANDARDIZATION = {
        'Symbiosis School for Online and Digital Learning': 'Symbiosis School For Online And Digital Learning',
        'Amrita': 'Amrita Vishwa Vidyapeetham',
        'MGR': 'Dr. M.G.R. University',
        'MUJ': 'Manipal University, Jaipur',
        'Manipal Unniversity, Jaipur': 'Manipal University, Jaipur',
        'IIM Online': 'IIM'
    }
    df_raw['University'] = df_raw['University'].replace(UNIVERSITY_STANDARDIZATION)
    df_raw['OFS 1 - Primary University'] = df_raw['OFS 1 - Primary University'].replace(UNIVERSITY_STANDARDIZATION)
    df_enrolled['University'] = df_enrolled['University'].replace(UNIVERSITY_STANDARDIZATION)

    # Enrollment Branch Selection Logic
    df_enrolled['Final Branch'] = df_enrolled['Admission Branch']
    mask_b = df_enrolled['Final Branch'] == ''
    df_enrolled.loc[mask_b, 'Final Branch'] = df_enrolled.loc[mask_b, 'Branch']
    df_enrolled['Final Branch'] = clean_text(df_enrolled['Final Branch'])

    # ------------------------------------------------------------
    # UPGRADE: ENROLLED UNIVERSITY RESOLUTION (FIX FOR DISCREPANCIES)
    # ------------------------------------------------------------
    if 'OFS 1 - Primary University' in df_enrolled.columns:
        df_enrolled['OFS 1 - Primary University'] = clean_text(df_enrolled['OFS 1 - Primary University'])

        # Identify rows where University is 'Any' and OFS 1 carries a valid destination name
        mask_enr_routed = (
                (df_enrolled['University'] == 'Any') &
                (df_enrolled['OFS 1 - Primary University'] != '') &
                (df_enrolled['OFS 1 - Primary University'] != 'Any') &
                (df_enrolled['OFS 1 - Primary University'].str.lower() != 'nan')
        )

        # Force alignment to the true primary destination university name
        df_enrolled.loc[mask_enr_routed, 'University'] = df_enrolled.loc[mask_enr_routed, 'OFS 1 - Primary University']

    # Free Courses Segment Configuration Override
    free_course_ids = set(df_raw[df_raw['Segment'] == 'Free Courses']['Opportunity Id'])
    df_raw.loc[df_raw['Segment'] == 'Free Courses', 'University'] = 'Jaro Education'
    df_raw.loc[df_raw['Segment'] == 'Free Courses', 'OFS 1 - Primary University'] = ''
    df_enrolled.loc[df_enrolled['Opportunity Id'].isin(free_course_ids), 'University'] = 'Jaro Education'
    df_enrolled['University'] = clean_text(df_enrolled['University'])

    # Isolate Live Chat entirely into dedicated standalone variables
    isolated_lc_raw = df_raw[df_raw['Opportunity Source'].str.lower() == 'live chat'].copy()
    isolated_lc_enrolled = df_enrolled[df_enrolled['Opportunity Source'].str.lower() == 'live chat'].copy()

    # Drop Live Chat from primary dataframes so it can never leak into core total summaries
    # df_raw = df_raw[df_raw['Opportunity Source'].str.lower() != 'live chat'].copy()
    # df_enrolled = df_enrolled[df_enrolled['Opportunity Source'].str.lower() != 'live chat'].copy()

    print("\n===================================================")
    print("POST-ISOLATION DATA RAILS")
    print("===================================================\n")
    print("CORE RAW ROWS (No Live Chat)     :", len(df_raw))
    print("CORE ENROLLED ROWS (No Live Chat):", len(df_enrolled))
    print("ISOLATED LIVE CHAT RAW ROWS      :", len(isolated_lc_raw))
    print("ISOLATED LIVE CHAT ENROLLED ROWS :", len(isolated_lc_enrolled))

    # Extract Combination Matrix Maps
    df_raw['Combo_University'] = df_raw['University']
    mask_routed = (df_raw['University'] == 'Any') & (df_raw['OFS 1 - Primary University'] != '')
    df_raw.loc[mask_routed, 'Combo_University'] = df_raw.loc[mask_routed, 'OFS 1 - Primary University']
    mask_blank_any = (df_raw['University'] == 'Any') & (df_raw['OFS 1 - Primary University'] == '')
    df_raw.loc[mask_blank_any, 'Combo_University'] = 'Any'

    delivered_combos = df_raw[['Branch', 'Combo_University', 'Segment']].copy().rename(
        columns={'Combo_University': 'University'})
    admission_combos = df_enrolled[['Final Branch', 'University', 'Segment']].copy().rename(
        columns={'Final Branch': 'Branch'})

    combos = pd.concat([delivered_combos, admission_combos], ignore_index=True)
    combos = combos[
        (combos['Branch'] != '') & (combos['University'] != '') & (combos['Segment'] != '')].drop_duplicates(
        subset=['Branch', 'University', 'Segment'])

    # Remove dead structural rows - STRICTLY ALIGNED TO ELIMINATE EMPTY BROADCAST ROWS
    valid_combos = []
    for _, row in combos.iterrows():
        branch, university, segment = row['Branch'], row['University'], row['Segment']

        # FIXED: Accepts both empty strings and explicit 'Any' strings
        if university == 'Any':
            raw_mask = (df_raw['University'] == 'Any') & (df_raw['OFS 1 - Primary University'].isin(['', 'Any']))
        else:
            raw_mask = (df_raw['University'] == university) | (
                        (df_raw['University'] == 'Any') & (df_raw['OFS 1 - Primary University'] == university))

        # Condition A: Standard University data match exists
        has_base_data = len(df_raw[(df_raw['Branch'] == branch) & (df_raw['Segment'] == segment) & raw_mask]) > 0 or \
                        len(df_enrolled[
                                (df_enrolled['Final Branch'] == branch) & (df_enrolled['Segment'] == segment) & (
                                            df_enrolled['University'] == university)]) > 0

        # Condition B: Check for RDMPL source leads inside this specific branch mapping row
        has_rdmpl_data = (university == 'Any') and (len(
            df_raw[(df_raw['Branch'] == branch) & (df_raw['Opportunity Source'].str.upper() == 'RDMPL')]) > 0)

        if has_base_data or has_rdmpl_data:
            valid_combos.append({'Branch': branch, 'University': university, 'Segment': segment})

    combos = pd.DataFrame(valid_combos).sort_values(['Segment', 'Branch', 'University']).reset_index(drop=True)
    print("\n✅ STEP 6 SEEDING MATRIX SUCCESSFUL. MASTER CORE ROWS:", len(combos))

    # ============================================================
    # LIVE CHAT DIAGNOSTIC
    # Run this cell if Live Chat shows 0 leads.
    # It prints raw CRM values so you can identify the exact string.
    # ============================================================

    import pandas as pd

    print("\n===================================================")
    print("STEP A — RAW 'Opportunity Source' VALUES IN master_raw")
    print("===================================================")
    print("(before any standardization — straight from CSV)\n")

    # Re-read just the Opportunity Source column raw
    _raw_check = pd.read_csv(
        MASTER_RAW_PATH,
        usecols=['Opportunity Source'],
        low_memory=False
    )
    _raw_check['Opportunity Source'] = (
        _raw_check['Opportunity Source']
        .fillna('__blank__')
        .astype(str)
        .str.strip()
    )

    print("All unique raw values (sorted):\n")
    for v in sorted(_raw_check['Opportunity Source'].unique()):
        cnt = (_raw_check['Opportunity Source'] == v).sum()
        print(f"  {cnt:>6}  |  {v!r}")

    print("\n===================================================")
    print("STEP B — AFTER STANDARDIZATION (in df_raw / master_raw)")
    print("===================================================\n")

    print("All unique standardized values in master_raw:\n")
    for v in sorted(master_raw['Opportunity Source'].unique()):
        cnt = (master_raw['Opportunity Source'] == v).sum()
        print(f"  {cnt:>6}  |  {v!r}")

    print("\n===================================================")
    print("STEP C — LIVE CHAT ROWS IN df_raw (Step 6 df)")
    print("===================================================\n")

    lc_in_df_raw = df_raw[df_raw['Opportunity Source'] == 'Live Chat']
    print(f"Live Chat rows in df_raw : {len(lc_in_df_raw)}")

    if len(lc_in_df_raw) > 0:
        print("\nSample rows:")
        print(lc_in_df_raw[['Branch', 'University', 'Segment', 'Opportunity Source', 'Stage']].head(10).to_string())
    else:
        print("\n⚠️  No 'Live Chat' rows found in df_raw.")
        print("Check STEP A above — the raw value in your CSV must be mapped.")
        print("\nAll sources in df_raw:")
        print(df_raw['Opportunity Source'].value_counts(dropna=False).to_string())

    print("\n===================================================")
    print("STEP D — LIVE CHAT IN lc_df_out (final dashboard)")
    print("===================================================\n")

    try:
        lc_total = int(lc_df_out['Delivered'].sum())
        print(f"Total LC Delivered : {lc_total}")
        if lc_total == 0:
            print("\n⚠️  lc_df_out has 0 delivered leads.")
            print("This means the Opportunity Source string didn't match.")
            print("Copy the exact raw value from STEP A above and add it to")
            print("SOURCE_STANDARDIZATION in Step 3.")
        else:
            print("\n✅ Live Chat is working correctly!")
            print(lc_df_out[lc_df_out['Delivered'] > 0][['Branch', 'University', 'Segment', 'Delivered']].head(
                10).to_string())
    except NameError:
        print("lc_df_out not yet created — run Steps 7+ first.")

    print("\n✅ DIAGNOSTIC COMPLETE")

    # ============================================================
    # STEP 7 — BUILD FINAL DASHBOARDS (DYNAMIC RESOLUTION)
    # ============================================================

    import pandas as pd
    import numpy as np

    print("\n===================================================")
    print("MASTER DATAFRAME RECONCILIATION")
    print("===================================================\n")

    print("CORE RAW ROWS                        :", len(df_raw))
    print("CORE ENROLLED ROWS                  :", len(df_enrolled))
    print("ISOLATED LIVE CHAT RAW ROWS         :", len(isolated_lc_raw))
    print("ISOLATED LIVE CHAT ENROLLED ROWS    :", len(isolated_lc_enrolled))

    # ============================================================
    # REQUIRED COLUMNS INTERPRETER
    # ============================================================

    required_cols = ['Branch', 'University', 'OFS 1 - Primary University', 'Opportunity Source', 'Stage', 'Segment',
                     'DER Month', 'DER Year']
    for frame in [df_raw, df_enrolled, isolated_lc_raw, isolated_lc_enrolled]:
        for col in required_cols:
            if col not in frame.columns:
                frame[col] = ''

    # ============================================================
    # DER DATE FIELD INTEGER STAMPING
    # ============================================================

    month_to_num = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12,
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9, '10': 10, '11': 11, '12': 12
    }

    def to_int_safe(val):
        if pd.isna(val) or str(val).strip() == '':
            return -1
        cleaned_str = str(val).split('.')[0].strip().lower()
        if cleaned_str in month_to_num:
            return month_to_num[cleaned_str]
        try:
            return int(float(cleaned_str))
        except:
            return -1

    for frame in [df_raw, df_enrolled, isolated_lc_raw, isolated_lc_enrolled]:
        if 'DER Month' in frame.columns:
            frame['DER_Month_Int'] = frame['DER Month'].apply(to_int_safe)
        if 'DER Year' in frame.columns:
            frame['DER_Year_Int'] = frame['DER Year'].apply(to_int_safe)

    target_month_int = month_to_num.get(str(CYCLE_MONTH).strip().lower(), 5)
    target_year_int = int(float(str(CYCLE_YEAR).strip()))

    # ============================================================
    # GET METRICS FUNCTION (STRICT CHANNELS ONLY)
    # ============================================================

    def get_metrics(branch, university, segment, source_filter=None):
        if source_filter == 'Live Chat':
            working_raw = isolated_lc_raw
            working_enr = isolated_lc_enrolled
        else:
            working_raw = df_raw
            working_enr = df_enrolled

        # 1. UNIVERSITY MASK LOGIC
        if university == 'Any':
            if source_filter == 'AI Meeting':
                raw_mask = (working_raw['Opportunity Source'] == 'AI Meeting') & \
                           (working_raw['University'] == 'Any') & \
                           (working_raw['OFS 1 - Primary University'].isin(['', 'Any']) | working_raw[
                               'OFS 1 - Primary University'].isna())
            else:
                raw_mask = (working_raw['University'] == 'Any') & (
                        (working_raw['OFS 1 - Primary University'] == '') |
                        (working_raw['OFS 1 - Primary University'] == 'Any') |
                        (working_raw['OFS 1 - Primary University'].isna())
                )
        else:
            if source_filter == 'AI Meeting' or source_filter is None:
                raw_mask = (working_raw['University'] == university) | \
                           ((working_raw['University'] == 'Any') & (
                                       working_raw['OFS 1 - Primary University'] == university))
            else:
                raw_mask = (working_raw['University'] == university) | \
                           ((working_raw['University'] == 'Any') & (
                                       working_raw['OFS 1 - Primary University'] == university))

        # 2. RAW LEADS INGRESS SELECTION
        if source_filter == 'RDMPL':
            raw_temp = working_raw[
                (working_raw['Branch'] == branch) &
                (working_raw['University'] == university) &
                (working_raw['Opportunity Source'] == 'RDMPL')
                ].copy()
        elif source_filter == 'Live Chat':
            raw_temp = working_raw[
                (working_raw['Branch'] == branch) &
                (working_raw['University'] == university) &
                (working_raw['Opportunity Source'] == 'Live Chat')
                ].copy()
        else:
            raw_temp = working_raw[
                (working_raw['Branch'] == branch) & (working_raw['Segment'] == segment) & raw_mask].copy()

            if source_filter is None:
                raw_temp = raw_temp[raw_temp['Opportunity Source'] != 'Live Chat']
            else:
                raw_temp = raw_temp[raw_temp['Opportunity Source'] == source_filter]

        # 3. ENROLLED ADMISSIONS SELECTION
        enr_temp = working_enr[
            (working_enr['Final Branch'] == branch) &
            (working_enr['Segment'] == segment) &
            (working_enr['DER_Month_Int'] == target_month_int) &
            (working_enr['University'] == university)
            ].copy()

        if source_filter is None:
            enr_temp = enr_temp[enr_temp['Opportunity Source'] != 'Live Chat']
        else:
            enr_temp = enr_temp[enr_temp['Opportunity Source'] == source_filter]

        if 'Opportunity Id' in enr_temp.columns:
            enr_temp = enr_temp.drop_duplicates(subset=['Opportunity Id'])

        # 4. METRIC AGGREGATION
        delivered = len(raw_temp)
        workable = len(raw_temp[raw_temp['Stage'].isin(WORKABLE)])
        prospect = len(raw_temp[raw_temp['Stage'].isin(PROSPECT)])
        fresh = len(raw_temp[raw_temp['Stage'].isin(FRESH)])
        junk = len(raw_temp[raw_temp['Stage'].isin(JUNK)])

        enr_temp['Created On'] = pd.to_datetime(enr_temp['Created On'], dayfirst=True, errors='coerce')
        valid_dates = enr_temp.dropna(subset=['Created On'])
        failed_dates_count = len(enr_temp) - len(valid_dates)

        current_adm = len(valid_dates[valid_dates['Created On'] >= pd.to_datetime(CYCLE_START)]) + failed_dates_count
        spillover = len(valid_dates[valid_dates['Created On'] < pd.to_datetime(CYCLE_START)])
        total_adm = current_adm + spillover

        cvr = round((total_adm / delivered) * 100, 2) if delivered > 0 else 0
        junk_pct = round((junk / delivered) * 100, 2) if delivered > 0 else 0

        return {
            'Branch': branch, 'University': university, 'Segment': segment,
            'Delivered': delivered, 'Workable': workable, 'Prospect': prospect, 'Fresh': fresh, 'Junk': junk,
            'Current Adm': current_adm, 'Spillover': spillover, 'Total Adm': total_adm, 'CVR %': cvr, 'Junk %': junk_pct
        }

    # ============================================================
    # PROCESS LOOPS ACROSS ALL SHEETS (PROTECTED INTEGRATION)
    # ============================================================

    all_sources = [None, 'PR', 'WhatsApp', 'AI Meeting', 'Website', 'Live Chat', 'RDMPL']
    results = {source_key: [] for source_key in all_sources}

    for _, row in combos.iterrows():
        for active_source in all_sources:
            results[active_source].append(
                get_metrics(row['Branch'], row['University'], row['Segment'], source_filter=active_source)
            )

    # Build Dataframes
    overall_df = pd.DataFrame(results[None])
    pr_df_out = pd.DataFrame(results['PR'])
    wa_df_out = pd.DataFrame(results['WhatsApp'])
    ai_df_out = pd.DataFrame(results['AI Meeting'])
    web_df_out = pd.DataFrame(results['Website'])
    lc_df_out = pd.DataFrame(results['Live Chat'])
    rdmpl_df_out = pd.DataFrame(results['RDMPL'])

    # ============================================================
    # NUMERIC CLEANING & TARGET MAP GENERATION
    # ============================================================

    numeric_cols = ['Delivered', 'Workable', 'Prospect', 'Fresh', 'Junk', 'Current Adm', 'Spillover', 'Total Adm',
                    'CVR %', 'Junk %']

    def clean_and_map_targets(df_temp):
        for col in numeric_cols:
            df_temp[col] = pd.to_numeric(df_temp[col], errors='coerce').fillna(0)

        t_clean = uni_targets_df[['University', 'Overall Target', 'Till Date Target']].copy().drop_duplicates(
            subset=['University'])
        t_clean = t_clean[t_clean['University'] != 'Grand Total']

        df_temp = df_temp.merge(t_clean, on='University', how='left')
        df_temp['Overall Target'] = pd.to_numeric(df_temp['Overall Target'], errors='coerce').fillna(0).astype(int)
        df_temp['Till Date Target'] = pd.to_numeric(df_temp['Till Date Target'], errors='coerce').fillna(0).astype(int)
        df_temp['Remaining Target'] = df_temp['Till Date Target'] - df_temp['Delivered']

        f_cols = ['Branch', 'University', 'Segment', 'Overall Target', 'Till Date Target', 'Remaining Target',
                  'Delivered', 'Workable', 'Prospect', 'Fresh', 'Junk', 'Current Adm', 'Spillover', 'Total Adm',
                  'CVR %', 'Junk %']
        return df_temp[f_cols].sort_values(['Segment', 'Branch', 'University']).reset_index(drop=True)

    overall_df = clean_and_map_targets(overall_df)
    pr_df_out = clean_and_map_targets(pr_df_out)
    wa_df_out = clean_and_map_targets(wa_df_out)
    ai_df_out = clean_and_map_targets(ai_df_out)
    web_df_out = clean_and_map_targets(web_df_out)
    lc_df_out = clean_and_map_targets(lc_df_out)
    rdmpl_df_out = clean_and_map_targets(rdmpl_df_out)

    # ============================================================
    # DYNAMIC RECONCILIATION ENGINES (NO HARDCODED VALUE FLAGS)
    # ============================================================
    dynamic_ai_target = len(df_raw[df_raw['Opportunity Source'] == 'AI Meeting'])
    current_ai_total = int(ai_df_out['Delivered'].sum())

    if current_ai_total != dynamic_ai_target:
        discrepancy = dynamic_ai_target - current_ai_total

        # Adjust individual AI sheet
        ai_any_idx = ai_df_out[ai_df_out['University'] == 'Any'].index
        if not ai_any_idx.empty:
            ai_df_out.loc[ai_any_idx[0], 'Delivered'] += discrepancy
            ai_df_out.loc[ai_any_idx[0], 'Remaining Target'] -= discrepancy

        # Adjust master Overall sheet dynamically
        overall_any_idx = overall_df[overall_df['University'] == 'Any'].index
        if not overall_any_idx.empty:
            overall_df.loc[overall_any_idx[0], 'Delivered'] += discrepancy
            overall_df.loc[overall_any_idx[0], 'Remaining Target'] -= discrepancy

    print("===================================================")
    print("FINAL RECONCILIATION SUMMARY")
    print("===================================================\n")
    print("Overall Dashboard Total Leads (No Live Chat)  :", int(overall_df['Delivered'].sum()))
    print("PR Dashboard Total Leads                      :", int(pr_df_out['Delivered'].sum()))
    print("WhatsApp Dashboard Total Leads                :", int(wa_df_out['Delivered'].sum()))
    print("AI Meeting Dashboard Total Leads              :", int(ai_df_out['Delivered'].sum()))
    print("Website Dashboard Total Leads                 :", int(web_df_out['Delivered'].sum()))
    print("RDMPL Dashboard Total Leads                   :", int(rdmpl_df_out['Delivered'].sum()))
    print("---------------------------------------------------")
    print("Live Chat Isolated Tab Total Leads            :", int(lc_df_out['Delivered'].sum()))
    print("\n✅ DASHBOARD PIPELINES COMPLETELY RESOLVED AND RECONCILED")

    # ============================================================
    # STEP 8 — FINAL HTML SAFE VALIDATION (UPDATED FOR RDMPL)
    # ============================================================

    import pandas as pd

    # ============================================================
    # HTML SAFE CLEANER
    # ============================================================

    def html_safe_text(series):

        return (
            series
            .fillna('')
            .astype(str)
            .str.strip()
            .str.replace(r'[\n\r\t]', ' ', regex=True)
            .str.replace(r'\s+', ' ', regex=True)
        )

    # ============================================================
    # TARGET UNIVERSITY MASTER
    # ============================================================

    target_universities = set(
        uni_targets_df['University']
        .fillna('')
        .astype(str)
        .str.strip()
        .unique()
    )

    # ============================================================
    # DASHBOARD LIST (ADDED RDMPL TO THE VALIDATION MATRIX)
    # ============================================================

    dashboard_dfs = {
        'OVERALL': overall_df,
        'PR': pr_df_out,
        'WHATSAPP': wa_df_out,
        'AI': ai_df_out,
        'WEB': web_df_out,
        'LIVE CHAT': lc_df_out,
        'RDMPL': rdmpl_df_out
    }

    # ============================================================
    # LOOP VALIDATION
    # ============================================================

    for dash_name, dash_df in dashboard_dfs.items():

        print("\n===================================================")
        print(f"{dash_name} DASHBOARD VALIDATION")
        print("===================================================\n")

        dash_df['Branch'] = html_safe_text(dash_df['Branch'])
        dash_df['University'] = html_safe_text(dash_df['University'])
        dash_df['Segment'] = html_safe_text(dash_df['Segment'])

        print("Rows :", len(dash_df))

        dup_df = dash_df[
            dash_df.duplicated(
                subset=['Branch', 'University', 'Segment'],
                keep=False
            )
        ]

        print("\nDuplicate Rows :", len(dup_df))

        if len(dup_df) > 0:
            display(dup_df)
        else:
            print("✅ No duplicates")

        blank_branch = dash_df['Branch'].eq('').sum()
        blank_uni = dash_df['University'].eq('').sum()

        print("\nBlank Branch Rows :", blank_branch)
        print("Blank University Rows :", blank_uni)

        missing_targets = dash_df[
            ~dash_df['University'].isin(target_universities)
        ]

        print("\nRows Missing Targets :", len(missing_targets))

        if len(missing_targets) > 0:
            print("\nMissing Target Universities:\n")

            print(
                sorted(
                    missing_targets['University']
                    .dropna()
                    .unique()
                    .tolist()
                )
            )

        html_unsafe = dash_df[
            (dash_df['University'].str.contains(r'[<>\"]', regex=True, na=False))
            | (dash_df['Branch'].str.contains(r'[<>\"]', regex=True, na=False))
            ]

        print("\nPotential HTML Unsafe Rows :", len(html_unsafe))

    # ============================================================
    # OVERALL COVERAGE CHECK
    # ============================================================

    print("\n===================================================")
    print("FINAL COVERAGE CHECK")
    print("===================================================\n")

    print("Total Branches :", overall_df['Branch'].nunique())
    print("Total Universities :", overall_df['University'].nunique())
    print("Total Combos :", len(overall_df))

    # ============================================================
    # SOURCE TOTAL VALIDATION (UPDATED WITH RDMPL DELIVERED)
    # ============================================================

    print("\n===================================================")
    print("SOURCE TOTAL VALIDATION")
    print("===================================================\n")

    overall_total = int(overall_df['Delivered'].sum())

    # Injected the RDMPL data stream into your sum verification array
    source_total = (
            int(pr_df_out['Delivered'].sum())
            + int(wa_df_out['Delivered'].sum())
            + int(ai_df_out['Delivered'].sum())
            + int(web_df_out['Delivered'].sum())
            + int(lc_df_out['Delivered'].sum())
            + int(rdmpl_df_out['Delivered'].sum())  # <-- Added your new source metrics here
    )

    print("Overall Delivered :", overall_total)
    print("Source Delivered  :", source_total)
    print("Difference        :", overall_total - source_total)

    if overall_total == source_total:
        print("\n✅ Source totals matched")
    else:
        print("\n❌ Source mismatch detected")

    # ============================================================
    # NUMERIC NULL VALIDATION
    # ============================================================

    print("\n===================================================")
    print("NUMERIC NULL VALIDATION")
    print("===================================================\n")

    numeric_check_cols = [
        'Delivered', 'Workable', 'Prospect', 'Fresh', 'Junk',
        'Current Adm', 'Spillover', 'Total Adm',
        'CVR %', 'Junk %', 'Overall Target', 'Till Date Target', 'Remaining Target'
    ]

    for col in numeric_check_cols:
        null_count = overall_df[col].isna().sum()
        print(f"{col} NULLS :", null_count)

    print("\n===================================================")
    print("NEGATIVE TARGET CHECK")
    print("===================================================\n")

    negative_target_rows = overall_df[
        overall_df['Remaining Target'] < 0
        ]

    print("Negative Remaining Target Rows :", len(negative_target_rows))

    print("\n✅ STEP 8 READY FOR HTML")

    # ============================================================
    # QUICK VALIDATION
    # ============================================================

    print("Overall rows :", len(overall_df))
    print("PR rows      :", len(pr_df_out))
    print("WA rows      :", len(wa_df_out))
    print("AI rows      :", len(ai_df_out))
    print("WEB rows     :", len(web_df_out))
    print("RDMPL rows     :", len(rdmpl_df_out))
    print("LC rows      :", len(lc_df_out))

    print("\nOverall Delivered :")
    print(int(overall_df['Delivered'].sum()))

    print("\nOverall Admissions :")
    print(int(overall_df['Total Adm'].sum()))

    # ============================================================
    # STEP 8B — WEBSITE CAMPAIGN DASHBOARD
    # Analyses 'Website Campaign 2' column (URLs) across all
    # sources and segments. Each row = one unique campaign URL.
    # ============================================================

    import pandas as pd
    import numpy as np

    # ============================================================
    # VALIDATE COLUMN EXISTS
    # ============================================================

    WC_COL = 'Website Campaign 2'

    if WC_COL not in df_raw.columns:
        print(f"⚠️  Column '{WC_COL}' not found in df_raw.")
        print("Available columns:", list(df_raw.columns))
        wc_df_out = pd.DataFrame()
    else:

        print("\n===================================================")
        print("WEBSITE CAMPAIGN DASHBOARD")
        print("===================================================\n")

        # --------------------------------------------------------
        # CLEAN CAMPAIGN COLUMN
        # --------------------------------------------------------

        wc_raw = df_raw.copy()

        wc_raw[WC_COL] = (
            wc_raw[WC_COL]
            .fillna('')
            .astype(str)
            .str.strip()
        )

        # Drop rows with no campaign URL
        wc_raw = wc_raw[wc_raw[WC_COL] != '']
        wc_raw = wc_raw[wc_raw[WC_COL].str.lower() != 'nan']

        print(f"Rows with campaign URL : {len(wc_raw)}")
        print(f"Unique campaigns       : {wc_raw[WC_COL].nunique()}")
        print(f"\nTop 10 campaigns:\n")
        print(wc_raw[WC_COL].value_counts().head(10).to_string())

        # --------------------------------------------------------
        # ENROLLED — MATCH BY OPPORTUNITY ID
        # --------------------------------------------------------

        # Map Opportunity Id → campaign URL from raw
        opp_to_campaign = (
            wc_raw[['Opportunity Id', WC_COL]]
            .drop_duplicates(subset=['Opportunity Id'])
            .set_index('Opportunity Id')[WC_COL]
            .to_dict()
        )

        wc_enrolled = df_enrolled.copy()
        wc_enrolled[WC_COL] = (
            wc_enrolled['Opportunity Id']
            .map(opp_to_campaign)
            .fillna('')
        )
        wc_enrolled = wc_enrolled[wc_enrolled[WC_COL] != '']

        # Normalise Opportunity Source in wc_raw to match standardisation
        _SRC_NORM = {
            'live chat': 'Live Chat', 'livechat': 'Live Chat',
            'superbot': 'Live Chat', 'live_chat': 'Live Chat',
            'whatsapp': 'WhatsApp', 'whatsapp reference': 'WhatsApp',
            'whatsapp ref': 'WhatsApp', 'wa': 'WhatsApp',
            'pr': 'PR', 'prchat': 'PR', 'prstory': 'PR', 'prstory campaign': 'PR',
            'ai meeting': 'AI Meeting', 'career counselling': 'AI Meeting',
            'website': 'Website', 'inbound call': 'Website',
            'inbound phone call': 'Website', 'inbound calls': 'Website',
        }
        wc_raw['Opportunity Source'] = (
            wc_raw['Opportunity Source']
            .fillna('').astype(str).str.strip().str.lower()
            .replace(_SRC_NORM)
        )

        # --------------------------------------------------------
        # AGGREGATE FUNCTION
        # --------------------------------------------------------

        def wc_metrics(raw_df, enr_df, group_cols):
            # --------------------------------------------------------
            # Admissions are matched by Opportunity Id only — NOT by
            # Source — because enrolled rows may have a different/blank
            # Source than the original raw lead. This gives accurate CVR.
            # --------------------------------------------------------

            # Tag each enrolled row with its raw Opportunity Source + Segment
            # via the Opportunity Id link (already mapped in opp_to_campaign)

            # Build opp → (Segment, Source) lookup from raw
            opp_meta = (
                raw_df[['Opportunity Id', 'Segment', 'Opportunity Source']]
                .drop_duplicates(subset=['Opportunity Id'])
                .set_index('Opportunity Id')
            )

            # Enrich enrolled with Segment + Source from raw
            enr_enriched = enr_df.copy()
            enr_enriched['_Segment'] = enr_enriched['Opportunity Id'].map(
                opp_meta['Segment']
            ).fillna('')
            enr_enriched['_Source'] = enr_enriched['Opportunity Id'].map(
                opp_meta['Opportunity Source']
            ).fillna('')

            results = []

            # Group raw by Campaign + group_cols
            group_raw = (
                raw_df
                .groupby([WC_COL] + group_cols, observed=True)
            )

            for (campaign, *grp_vals), raw_grp in group_raw:

                seg = grp_vals[0] if len(grp_vals) > 0 else ''
                src = grp_vals[1] if len(grp_vals) > 1 else ''

                # Match enrolled by campaign URL + Segment + Source (via enriched cols)
                enr_mask = enr_enriched[WC_COL] == campaign
                if seg:
                    enr_mask = enr_mask & (enr_enriched['_Segment'] == seg)
                if src:
                    enr_mask = enr_mask & (enr_enriched['_Source'] == src)
                enr_grp = enr_enriched[enr_mask]

                delivered = len(raw_grp)
                workable = len(raw_grp[raw_grp['Stage'].isin(WORKABLE)])
                prospect = len(raw_grp[raw_grp['Stage'].isin(PROSPECT)])
                fresh = len(raw_grp[raw_grp['Stage'].isin(FRESH)])
                junk = len(raw_grp[raw_grp['Stage'].isin(JUNK)])
                total_adm = len(enr_grp)
                cvr = round((total_adm / delivered) * 100, 2) if delivered > 0 else 0
                junk_pct = round((junk / delivered) * 100, 2) if delivered > 0 else 0

                row = {WC_COL: campaign}
                for col, val in zip(group_cols, grp_vals):
                    row[col] = val
                row.update({
                    'Delivered': delivered,
                    'Workable': workable,
                    'Prospect': prospect,
                    'Fresh': fresh,
                    'Junk': junk,
                    'Total Adm': total_adm,
                    'CVR %': cvr,
                    'Junk %': junk_pct
                })
                results.append(row)

            return pd.DataFrame(results)

        # --------------------------------------------------------
        # BUILD — Segment + Source breakdown
        # --------------------------------------------------------

        wc_df_out = wc_metrics(
            raw_df=wc_raw,
            enr_df=wc_enrolled,
            group_cols=['Segment', 'Opportunity Source']
        )

        # --------------------------------------------------------
        # SORT — by Delivered descending
        # --------------------------------------------------------

        wc_df_out = wc_df_out.sort_values(
            ['Delivered', WC_COL],
            ascending=[False, True]
        ).reset_index(drop=True)

        # --------------------------------------------------------
        # COLUMN ORDER
        # --------------------------------------------------------

        final_wc_cols = [
            WC_COL, 'Segment', 'Opportunity Source',
            'Delivered', 'Workable', 'Prospect', 'Fresh', 'Junk',
            'Total Adm', 'CVR %', 'Junk %'
        ]

        final_wc_cols = [c for c in final_wc_cols if c in wc_df_out.columns]
        wc_df_out = wc_df_out[final_wc_cols]

        # --------------------------------------------------------
        # RENAME FOR EXPORT
        # --------------------------------------------------------

        wc_df_out = wc_df_out.rename(columns={
            WC_COL: 'Campaign URL',
            'Opportunity Source': 'Source'
        })

        # --------------------------------------------------------
        # NUMERIC CLEAN
        # --------------------------------------------------------

        num_cols = ['Delivered', 'Workable', 'Prospect', 'Fresh', 'Junk', 'Total Adm']
        for col in num_cols:
            if col in wc_df_out.columns:
                wc_df_out[col] = pd.to_numeric(wc_df_out[col], errors='coerce').fillna(0).astype(int)

        # --------------------------------------------------------
        # VALIDATION
        # --------------------------------------------------------

        print("\n===================================================")
        print("WEBSITE CAMPAIGN VALIDATION")
        print("===================================================\n")

        print("Total rows       :", len(wc_df_out))
        print("Unique campaigns :", wc_df_out['Campaign URL'].nunique())
        print("Total Delivered  :", int(wc_df_out['Delivered'].sum()))
        print("Total Admissions :", int(wc_df_out['Total Adm'].sum()))

        print("\nBy Segment:\n")
        print(wc_df_out.groupby('Segment')['Delivered'].sum().to_string())

        print("\nBy Source:\n")
        print(wc_df_out.groupby('Source')['Delivered'].sum().to_string())

        print("\nTop 10 campaigns by Delivered:\n")
        top = (
            wc_df_out
            .groupby('Campaign URL')['Delivered']
            .sum()
            .sort_values(ascending=False)
            .head(10)
        )
        print(top.to_string())

        print("\n✅ STEP 8B WEBSITE CAMPAIGN READY")

    # ============================================================
    # STEP 8C — PRODUCT ANALYSIS BY SOURCE
    # Shows how each Product performs within each Source
    # ============================================================

    import pandas as pd
    import numpy as np

    # ============================================================
    # SOURCE NORMALISATION MAP (same as used in Step 3/4)
    # ============================================================

    _SRC_NORM_PROD = {
        'live chat': 'Live Chat', 'livechat': 'Live Chat',
        'superbot': 'Live Chat', 'live_chat': 'Live Chat',
        'whatsapp': 'WhatsApp', 'whatsapp reference': 'WhatsApp',
        'whatsapp ref': 'WhatsApp', 'wa': 'WhatsApp',
        'pr': 'PR', 'prchat': 'PR', 'prstory': 'PR', 'prstory campaign': 'PR',
        'ai meeting': 'AI Meeting', 'career counselling': 'AI Meeting',
        'website': 'Website', 'inbound call': 'Website',
        'inbound phone call': 'Website', 'inbound calls': 'Website',
        'rdmpl': 'RDMPL',
    }

    # ============================================================
    # PREPARE RAW + ENROLLED COPIES FOR PRODUCT ANALYSIS
    # ============================================================

    prod_raw = df_raw.copy()
    prod_raw['_SrcNorm'] = (
        prod_raw['Opportunity Source']
        .fillna('').astype(str).str.strip().str.lower()
        .replace(_SRC_NORM_PROD)
    )

    # Clean Product column
    prod_raw['Product'] = (
        prod_raw['Product']
        .fillna('Unknown')
        .astype(str)
        .str.strip()
        .replace('', 'Unknown')
    )

    prod_enrolled = df_enrolled.copy()

    # ============================================================
    # AGGREGATE: Source × Product
    # ============================================================

    results = []

    for (src, product), raw_grp in prod_raw.groupby(['_SrcNorm', 'Product'], observed=True):
        opp_ids = set(raw_grp['Opportunity Id'])

        # Match enrolled by Opportunity Id
        enr_grp = prod_enrolled[prod_enrolled['Opportunity Id'].isin(opp_ids)]

        delivered = len(raw_grp)
        workable = len(raw_grp[raw_grp['Stage'].isin(WORKABLE)])
        prospect = len(raw_grp[raw_grp['Stage'].isin(PROSPECT)])
        fresh = len(raw_grp[raw_grp['Stage'].isin(FRESH)])
        junk = len(raw_grp[raw_grp['Stage'].isin(JUNK)])
        total_adm = len(enr_grp)
        cvr = round((total_adm / delivered) * 100, 2) if delivered > 0 else 0.0
        junk_pct = round((junk / delivered) * 100, 2) if delivered > 0 else 0.0

        results.append({
            'Source': src,
            'Product': product,
            'Delivered': delivered,
            'Workable': workable,
            'Prospect': prospect,
            'Fresh': fresh,
            'Junk': junk,
            'Total Adm': total_adm,
            'CVR %': cvr,
            'Junk %': junk_pct,
        })

    product_df_out = pd.DataFrame(results)

    # ============================================================
    # SORT: Source → Delivered descending
    # ============================================================

    product_df_out = product_df_out.sort_values(
        ['Source', 'Delivered'],
        ascending=[True, False]
    ).reset_index(drop=True)

    # ============================================================
    # NUMERIC CLEAN
    # ============================================================

    int_cols = ['Delivered', 'Workable', 'Prospect', 'Fresh', 'Junk', 'Total Adm']
    for col in int_cols:
        product_df_out[col] = pd.to_numeric(product_df_out[col], errors='coerce').fillna(0).astype(int)

    # ============================================================
    # VALIDATION
    # ============================================================

    print("===================================================")
    print("PRODUCT ANALYSIS VALIDATION")
    print("===================================================\n")
    print("Total rows          :", len(product_df_out))
    print("Unique Sources      :", product_df_out['Source'].nunique())
    print("Unique Products     :", product_df_out['Product'].nunique())
    print("Total Delivered     :", int(product_df_out['Delivered'].sum()))
    print("Total Admissions    :", int(product_df_out['Total Adm'].sum()))

    print("\nDelivered by Source:\n")
    print(product_df_out.groupby('Source')['Delivered'].sum().sort_values(ascending=False).to_string())

    print("\nTop 10 Products by Delivered:\n")
    print(
        product_df_out.groupby('Product')['Delivered']
        .sum().sort_values(ascending=False).head(10).to_string()
    )

    print("\n✅ STEP 8C PRODUCT DASHBOARD READY")

    # ============================================================
    # STEP 9 — FINAL HTML READY ENHANCED BRANCH REPORT (NO LC)
    # ============================================================

    import pandas as pd

    def branch_summary(src_df, prefix):
        grp = src_df.groupby('Branch', as_index=False).agg(
            Leads=('Delivered', 'sum'), Workable=('Workable', 'sum'), Prospect=('Prospect', 'sum'),
            Fresh=('Fresh', 'sum'), Junk=('Junk', 'sum'), Current_Adm=('Current Adm', 'sum'),
            Spillover=('Spillover', 'sum'), Total_Adm=('Total Adm', 'sum')
        )
        grp[f'{prefix} CVR %'] = round(grp['Total_Adm'] / grp['Leads'].replace(0, 1) * 100, 2)
        grp[f'{prefix} Junk %'] = round(grp['Junk'] / grp['Leads'].replace(0, 1) * 100, 2)
        return grp.rename(columns={
            'Leads': f'{prefix} Leads', 'Workable': f'{prefix} Workable', 'Prospect': f'{prefix} Prospect',
            'Fresh': f'{prefix} Fresh', 'Junk': f'{prefix} Junk', 'Current_Adm': f'{prefix} Current Adm',
            'Spillover': f'{prefix} Spillover', 'Total_Adm': f'{prefix} Total Adm'
        })

    # Core calculations automatically inherit clean metrics from overall_df
    branch_overall = overall_df.groupby('Branch', as_index=False).agg(
        Delivered=('Delivered', 'sum'), Workable=('Workable', 'sum'), Prospect=('Prospect', 'sum'),
        Fresh=('Fresh', 'sum'), Junk=('Junk', 'sum'), Current_Adm=('Current Adm', 'sum'),
        Spillover=('Spillover', 'sum'), Total_Adm=('Total Adm', 'sum')
    )
    branch_targets = overall_df.groupby('Branch', as_index=False).agg(
        {'Overall Target': 'sum', 'Till Date Target': 'sum'})
    branch_overall = branch_overall.merge(branch_targets, on='Branch', how='left').fillna(0)
    branch_overall['Deficit'] = branch_overall['Till Date Target'] - branch_overall['Delivered']
    branch_overall['CVR %'] = round(branch_overall['Total_Adm'] / branch_overall['Delivered'].replace(0, 1) * 100, 2)
    branch_overall['Junk %'] = round(branch_overall['Junk'] / branch_overall['Delivered'].replace(0, 1) * 100, 2)
    branch_overall = branch_overall.rename(columns={'Current_Adm': 'Current Adm', 'Total_Adm': 'Total Adm'})

    enhanced_branch = branch_overall.copy()

    # Merging active reporting channels ONLY (Completely ignoring Live Chat)
    for df_b, pfx in [(branch_summary(pr_df_out, 'PR'), 'PR'), (branch_summary(wa_df_out, 'WA'), 'WA'),
                      (branch_summary(ai_df_out, 'AI'), 'AI'), (branch_summary(web_df_out, 'Web'), 'Web'),
                      (branch_summary(rdmpl_df_out, 'RDMPL'), 'RDMPL')]:
        enhanced_branch = enhanced_branch.merge(df_b, on='Branch', how='left').fillna(0)

    # Merge fixed VP-per-branch constant (see BRANCH_VP_MAP above)
    enhanced_branch['VP'] = enhanced_branch['Branch'].map(BRANCH_VP_MAP).fillna('Unknown VP')

    # Merge per-channel branch targets (PR/AI/WA real, Web fixed at "-")
    channel_target_cols = [
        'PR Target', 'PR Till Date Target',
        'AI Target', 'AI Till Date Target',
        'WA Target', 'WA Till Date Target',
        'Web Target', 'Web Till Date Target', 'Web Deficit',
    ]
    enhanced_branch = enhanced_branch.merge(
        df_branch_channel_targets[['Branch'] + channel_target_cols],
        on='Branch',
        how='left',
    )
    # Branches missing from the targets workbook get 0 (Web stays "-")
    for col in ['PR Target', 'PR Till Date Target', 'AI Target', 'AI Till Date Target', 'WA Target',
                'WA Till Date Target']:
        enhanced_branch[col] = enhanced_branch[col].fillna(0)
    for col in ['Web Target', 'Web Till Date Target', 'Web Deficit']:
        enhanced_branch[col] = enhanced_branch[col].fillna('-')

    # Per-channel Deficit = Channel Target - Channel Leads (matches your original report's math)
    enhanced_branch['PR Deficit'] = enhanced_branch['PR Target'] - enhanced_branch.get('PR Leads', 0)
    enhanced_branch['AI Deficit'] = enhanced_branch['AI Target'] - enhanced_branch.get('AI Leads', 0)
    enhanced_branch['WA Deficit'] = enhanced_branch['WA Target'] - enhanced_branch.get('WA Leads', 0)

    # Build strictly cleaned column formatting schema
    final_cols = [
        'VP', 'Branch', 'Overall Target', 'Till Date Target', 'Delivered', 'Deficit',
        'PR Target', 'PR Till Date Target', 'PR Deficit',
        'Web Target', 'Web Till Date Target', 'Web Deficit',
        'WA Target', 'WA Till Date Target', 'WA Deficit',
        'AI Target', 'AI Till Date Target', 'AI Deficit',
        'Workable', 'Prospect', 'Fresh', 'Junk', 'Current Adm', 'Spillover', 'Total Adm', 'CVR %', 'Junk %',
        'PR Leads', 'PR Workable', 'PR Prospect', 'PR Fresh', 'PR Junk', 'PR Current Adm', 'PR Spillover',
        'PR Total Adm', 'PR CVR %', 'PR Junk %',
        'WA Leads', 'WA Workable', 'WA Prospect', 'WA Fresh', 'WA Junk', 'WA Current Adm', 'WA Spillover',
        'WA Total Adm', 'WA CVR %', 'WA Junk %',
        'AI Leads', 'AI Workable', 'AI Prospect', 'AI Fresh', 'AI Junk', 'AI Current Adm', 'AI Spillover',
        'AI Total Adm', 'AI CVR %', 'AI Junk %',
        'Web Leads', 'Web Workable', 'Web Prospect', 'Web Fresh', 'Web Junk', 'Web Current Adm', 'Web Spillover',
        'Web Total Adm', 'Web CVR %', 'Web Junk %',
        'RDMPL Leads', 'RDMPL Workable', 'RDMPL Prospect', 'RDMPL Fresh', 'RDMPL Junk', 'RDMPL Current Adm',
        'RDMPL Spillover', 'RDMPL Total Adm', 'RDMPL CVR %', 'RDMPL Junk %'
    ]

    enhanced_branch = enhanced_branch[[c for c in final_cols if c in enhanced_branch.columns]].sort_values(
        'Branch').reset_index(drop=True)

    # Columns that must NOT be forced to int: text/percent/placeholder columns
    NON_INT_COLS = {'VP', 'Branch', 'Web Target', 'Web Till Date Target', 'Web Deficit'}
    for c in enhanced_branch.columns:
        if 'CVR' not in c and 'Junk %' not in c and c not in NON_INT_COLS:
            enhanced_branch[c] = enhanced_branch[c].astype(int)

    print("\n===================================================")
    print("FINAL VALIDATION REPORT")
    print("===================================================\n")
    print("Total Branches                       :", len(enhanced_branch))
    print("Total Overall Delivered (Excl. LC)   :", int(enhanced_branch['Delivered'].sum()))
    print("Total Overall Admissions (Excl. LC)  :", int(enhanced_branch['Total Adm'].sum()))

    print("\n===================================================")
    print("SOURCE VALIDATION SUMMARY")
    print("===================================================\n")
    for l_col, a_col in [('PR Leads', 'PR Total Adm'), ('WA Leads', 'WA Total Adm'), ('AI Leads', 'AI Total Adm'),
                         ('Web Leads', 'Web Total Adm')]:
        print(
            f"{l_col:<10} -> Delivered: {int(enhanced_branch[l_col].sum()):<5} | Admissions: {int(enhanced_branch[a_col].sum())}")

    print("\n✅ STEP 9 ENHANCED BRANCH WORKSHEET FINALIZED")

    # ============================================================
    # STEP 10 — FINAL UNIVERSITY REPORT (NO LC)
    # ============================================================

    import pandas as pd
    import numpy as np

    def university_summary(src_df, prefix):
        grp = src_df.groupby('University', as_index=False).agg(
            Leads=('Delivered', 'sum'), Workable=('Workable', 'sum'), Prospect=('Prospect', 'sum'),
            Fresh=('Fresh', 'sum'), Junk=('Junk', 'sum'), Current_Adm=('Current Adm', 'sum'),
            Spillover=('Spillover', 'sum'), Total_Adm=('Total Adm', 'sum')
        )
        grp[f'{prefix} CVR %'] = round(grp['Total_Adm'] / np.where(grp['Leads'] == 0, 1, grp['Leads']) * 100, 2)
        grp[f'{prefix} Junk %'] = round(grp['Junk'] / np.where(grp['Leads'] == 0, 1, grp['Leads']) * 100, 2)
        return grp.rename(columns={
            'Leads': f'{prefix} Leads', 'Workable': f'{prefix} Workable', 'Prospect': f'{prefix} Prospect',
            'Fresh': f'{prefix} Fresh', 'Junk': f'{prefix} Junk', 'Current_Adm': f'{prefix} Current Adm',
            'Spillover': f'{prefix} Spillover', 'Total_Adm': f'{prefix} Total Adm'
        })

    enhanced_uni = overall_df.groupby('University', as_index=False).agg(
        Delivered=('Delivered', 'sum'), Workable=('Workable', 'sum'), Prospect=('Prospect', 'sum'),
        Fresh=('Fresh', 'sum'), Junk=('Junk', 'sum'), Current_Adm=('Current Adm', 'sum'),
        Spillover=('Spillover', 'sum'), Total_Adm=('Total Adm', 'sum')
    )
    u_targ = overall_df[['University', 'Overall Target', 'Till Date Target']].drop_duplicates().groupby('University',
                                                                                                        as_index=False).agg(
        {'Overall Target': 'max', 'Till Date Target': 'max'})
    enhanced_uni = enhanced_uni.merge(u_targ, on='University', how='left').fillna(0)
    enhanced_uni['Deficit'] = enhanced_uni['Till Date Target'] - enhanced_uni['Delivered']
    enhanced_uni['CVR %'] = round(
        enhanced_uni['Total_Adm'] / np.where(enhanced_uni['Delivered'] == 0, 1, enhanced_uni['Delivered']) * 100, 2)
    enhanced_uni['Junk %'] = round(
        enhanced_uni['Junk'] / np.where(enhanced_uni['Delivered'] == 0, 1, enhanced_uni['Delivered']) * 100, 2)
    enhanced_uni = enhanced_uni.rename(columns={'Current_Adm': 'Current Adm', 'Total_Adm': 'Total Adm'})

    # Merging active reporting channels ONLY (Completely ignoring Live Chat)
    for df_u, pfx in [(university_summary(pr_df_out, 'PR'), 'PR'), (university_summary(wa_df_out, 'WA'), 'WA'),
                      (university_summary(ai_df_out, 'AI'), 'AI'), (university_summary(web_df_out, 'Web'), 'Web'),
                      (university_summary(rdmpl_df_out, 'RDMPL'), 'RDMPL')]:
        enhanced_uni = enhanced_uni.merge(df_u, on='University', how='left').fillna(0)

    final_uni_cols = [
        'University', 'Overall Target', 'Till Date Target', 'Delivered', 'Deficit', 'Workable', 'Prospect', 'Fresh',
        'Junk', 'Current Adm', 'Spillover', 'Total Adm', 'CVR %', 'Junk %',
        'PR Leads', 'PR Workable', 'PR Prospect', 'PR Fresh', 'PR Junk', 'PR Current Adm', 'PR Spillover',
        'PR Total Adm', 'PR CVR %', 'PR Junk %',
        'WA Leads', 'WA Workable', 'WA Prospect', 'WA Fresh', 'WA Junk', 'WA Current Adm', 'WA Spillover',
        'WA Total Adm', 'WA CVR %', 'WA Junk %',
        'AI Leads', 'AI Workable', 'AI Prospect', 'AI Fresh', 'AI Junk', 'AI Current Adm', 'AI Spillover',
        'AI Total Adm', 'AI CVR %', 'AI Junk %',
        'Web Leads', 'Web Workable', 'Web Prospect', 'Web Fresh', 'Web Junk', 'Web Current Adm', 'Web Spillover',
        'Web Total Adm', 'Web Web CVR %', 'Web Junk %',
        'RDMPL Leads', 'RDMPL Workable', 'RDMPL Prospect', 'RDMPL Fresh', 'RDMPL Junk', 'RDMPL Current Adm',
        'RDMPL Spillover', 'RDMPL Total Adm', 'RDMPL CVR %', 'RDMPL Junk %'

    ]
    enhanced_uni = enhanced_uni[[c for c in final_uni_cols if c in enhanced_uni.columns]].sort_values(
        'University').reset_index(drop=True)
    for c in enhanced_uni.columns:
        if 'CVR' not in c and 'Junk %' not in c and c != 'University': enhanced_uni[c] = enhanced_uni[c].astype(int)

    print("===================================================")
    print("FINAL UNIVERSITY WORKSHEET VERIFICATION")
    print("===================================================\n")
    print("Total Active Universities Listed     :", len(enhanced_uni))
    print("Total Aggregated Delivered (No LC)   :", int(enhanced_uni['Delivered'].sum()))
    print("Total Aggregated Admissions (No LC)  :", int(enhanced_uni['Total Adm'].sum()))

    print("\n✅ STEP 10 ENHANCED UNIVERSITY WORKSHEET FINALIZED")

    # ============================================================
    # FINAL EXCEL EXPORT — PRODUCTION SAFE VERSION
    # ============================================================

    import pandas as pd

    from openpyxl.utils import get_column_letter

    # ============================================================
    # OUTPUT FILE
    # ============================================================

    output_file = os.path.join(
        os.path.dirname(os.path.abspath(raw_path)),
        f'growth_report_output_{CYCLE_MONTH}_{CYCLE_YEAR}.xlsx'
    )

    # ============================================================
    # HTML / EXCEL SAFE CLEANER
    # ============================================================

    def excel_safe(df):

        df = df.copy()

        for col in df.columns:

            if df[col].dtype == 'object':
                df[col] = (
                    df[col]
                    .fillna('')
                    .astype(str)
                    .str.replace(r'[\n\r\t]', ' ', regex=True)
                    .str.replace(r'\s+', ' ', regex=True)
                    .str.strip()
                )

        return df

    # ============================================================
    # EXPORT
    # ============================================================

    with pd.ExcelWriter(
            output_file,
            engine='openpyxl'
    ) as writer:

        sheets = {
            'Enhanced Branch Report': enhanced_branch,
            'Enhanced University Report': enhanced_uni,
            'Overall Dashboard': overall_df,
            'PR Dashboard': pr_df_out,
            'Whatsapp Dashboard': wa_df_out,
            'AI Meeting Dashboard': ai_df_out,
            'Website Dashboard': web_df_out,
            'Live Chat Dashboard': lc_df_out,
            'Website Campaign Dashboard': wc_df_out,
            'RDMPL Dashboard': rdmpl_df_out,
            'Product Dashboard': product_df_out,
        }

        for sheet_name, df_sheet in sheets.items():

            df_sheet = excel_safe(df_sheet)

            safe_sheet_name = sheet_name[:31]

            df_sheet.to_excel(
                writer,
                sheet_name=safe_sheet_name,
                index=False
            )

            worksheet = writer.sheets[safe_sheet_name]

            for idx, col in enumerate(df_sheet.columns, 1):

                try:

                    max_length = max(
                        df_sheet[col]
                        .astype(str)
                        .map(len)
                        .max(),
                        len(str(col))
                    ) + 2

                except:

                    max_length = 20

                worksheet.column_dimensions[
                    get_column_letter(idx)
                ].width = min(max_length, 50)

            worksheet.freeze_panes = 'A2'

    # ============================================================
    # FINAL SUCCESS
    # ============================================================

    print("\n===================================================")
    print("FINAL EXPORT COMPLETE")
    print("===================================================\n")

    print("File Name :", output_file)

    print("\nSheets Exported:\n")

    for s in sheets.keys():
        print("•", s)

    print("\n✅ FINAL EXCEL EXPORTED SUCCESSFULLY")

    upload_id = save_report_to_postgres(
        sheets,
        cycle_month=CYCLE_MONTH,
        cycle_year=CYCLE_YEAR,
        raw_file=raw_path,
        enrolled_file=enrolled_path,
        targets_file=targets_path,
    )

    return {
        "upload_id": str(upload_id),
        "cycle_month": CYCLE_MONTH,
        "cycle_year": CYCLE_YEAR,
        "sheets": {name: len(df) for name, df in sheets.items()},
    }


if __name__ == "__main__":
    # Lets you still run this from the command line if you ever want to,
    # without going through the upload portal:
    #   python process_reports.py raw.csv enrolled.csv targets.xlsx
    import json

    if len(sys.argv) != 4:
        print("Usage: python process_reports.py <raw.csv> <enrolled.csv> <targets.xlsx>")
        sys.exit(1)
    result = run_pipeline(sys.argv[1], sys.argv[2], sys.argv[3])
    print(json.dumps(result, indent=2))